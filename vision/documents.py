"""
CognitiveOC v3 — Document Parser
==================================
Parses any supported file type into (location, text) pairs.
Called by: perception/perception.py, retrieval/rag.py

Supported:  .txt .md .py .json .log .yaml .toml .csv .pdf .docx .xlsx .png .jpg .jpeg
Return:     list[tuple[str, str]]  — (location, text)
            location: 'text' | 'page N' | 'sheet:name' | 'image' | 'image_meta' | 'error'
"""

from __future__ import annotations
import csv
import re
from pathlib import Path


def parse_file(path) -> list[tuple[str, str]]:
    """Parse any supported file. Returns [(location, text), ...]."""
    p   = Path(path)
    ext = p.suffix.lower()

    # ── Plain text / code / data ──────────────────────────────────────
    if ext in {'.txt', '.md', '.py', '.json', '.log', '.yaml', '.yml', '.toml'}:
        try:
            return [('text', p.read_text(encoding='utf-8', errors='ignore'))]
        except Exception as e:
            return [('error', str(e))]

    # ── CSV ───────────────────────────────────────────────────────────
    if ext == '.csv':
        try:
            with p.open(newline='', encoding='utf-8', errors='ignore') as f:
                rows = list(csv.reader(f))
            if not rows:
                return [('table', '')]
            header = ', '.join(rows[0])
            body   = '\n'.join(', '.join(r) for r in rows[1:])
            return [('table', f'Columns: {header}\n{body}')]
        except Exception as e:
            return [('error', f'CSV parse error: {e}')]

    # ── PDF ───────────────────────────────────────────────────────────
    if ext == '.pdf':
        try:
            from pypdf import PdfReader
        except ImportError:
            return [('error', 'PDF support requires: pip install pypdf')]
        try:
            reader = PdfReader(str(p))
            return [(f'page {i+1}', pg.extract_text() or '')
                    for i, pg in enumerate(reader.pages)]
        except Exception as e:
            return [('error', f'PDF read error: {e}')]

    # ── DOCX ──────────────────────────────────────────────────────────
    if ext == '.docx':
        try:
            import docx
        except ImportError:
            return [('error', 'DOCX support requires: pip install python-docx')]
        try:
            doc      = docx.Document(str(p))
            sections: list[tuple[str, str]] = []
            heading, buf = 'document', []
            for par in doc.paragraphs:
                if par.style.name.startswith('Heading'):
                    if buf:
                        sections.append((heading, '\n'.join(buf)))
                    heading, buf = par.text or 'section', []
                elif par.text.strip():
                    buf.append(par.text)
            if buf:
                sections.append((heading, '\n'.join(buf)))
            return sections or [('document', '')]
        except Exception as e:
            return [('error', f'DOCX read error: {e}')]

    # ── XLSX ──────────────────────────────────────────────────────────
    if ext in {'.xlsx', '.xls'}:
        try:
            import openpyxl
        except ImportError:
            return [('error', 'XLSX support requires: pip install openpyxl')]
        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            results = []
            for name in wb.sheetnames:
                ws   = wb[name]
                rows = [[str(c.value or '') for c in row] for row in ws.iter_rows()]
                if not rows:
                    continue
                header = ', '.join(rows[0])
                body   = '\n'.join(', '.join(r) for r in rows[1:50])
                results.append((f'sheet:{name}', f'Columns: {header}\n{body}'))
            wb.close()
            return results or [('error', 'empty workbook')]
        except Exception as e:
            return [('error', f'XLSX read error: {e}')]

    # ── Images (OCR) ──────────────────────────────────────────────────
    if ext in {'.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.webp'}:
        try:
            from vision.ocr import analyze_image
            return analyze_image(str(p))
        except Exception as e:
            return [('error', f'OCR error: {e}')]

    return [('error', f'Unsupported file type: {ext}')]


def extract_text(path) -> str:
    """Convenience: return all text from a file as a single string."""
    parts = [_clean(t) for loc, t in parse_file(path)
             if loc != 'error' and t.strip()]
    return '\n\n'.join(parts)


def compare_documents(path_a: str, path_b: str) -> dict:
    """Jaccard similarity between two documents."""
    ta = extract_text(path_a)
    tb = extract_text(path_b)

    def tok(t):
        words = set(re.findall(r'[a-z0-9]+', t.lower()))
        return {w for w in words if len(w) > 2}

    ta_t, tb_t = tok(ta), tok(tb)
    shared = ta_t & tb_t
    union  = ta_t | tb_t
    return {
        'path_a':       path_a,
        'path_b':       path_b,
        'only_in_a':    sorted(ta_t - tb_t)[:30],
        'only_in_b':    sorted(tb_t - ta_t)[:30],
        'shared_count': len(shared),
        'jaccard':      round(len(shared) / max(len(union), 1), 4),
        'len_a_chars':  len(ta),
        'len_b_chars':  len(tb),
    }


def _clean(text: str) -> str:
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'[ \t]{2,}', ' ', text)
    return text.strip()
